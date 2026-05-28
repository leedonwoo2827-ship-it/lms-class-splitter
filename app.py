# -*- coding: utf-8 -*-
"""LMS 입과명단 분반기 (tkinter GUI).

여러 회사의 LMS 입과 엑셀을 입력받아 동일 과정끼리 모은 뒤
지정된 정원(기본 50명) 기준으로 N개 반으로 잘라 파일을 출력합니다.
"""
import os
import re
import threading
from collections import OrderedDict
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


SAFE_NAME_RE = re.compile(r'[\\/:*?"<>|]')


def sanitize_filename(s: str) -> str:
    return SAFE_NAME_RE.sub("_", str(s)).strip() or "미지정"


def read_file(path: str, course_col_name: str):
    """파일을 읽어 (headers, course_col_index_1based, data_rows) 반환."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        wb.close()
        return None, None, []

    col_idx = None
    name_norm = course_col_name.strip()
    for i, h in enumerate(headers, start=1):
        if h is not None and str(h).strip() == name_norm:
            col_idx = i
            break

    if col_idx is None:
        wb.close()
        return headers, None, []

    data = [list(r) for r in rows_iter if any(v is not None for v in r)]
    wb.close()
    return headers, col_idx, data


def write_xlsx(path: str, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "명단"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for ri, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=c, value=v)
            cell.alignment = center
            cell.border = border

    for c, h in enumerate(headers, start=1):
        max_len = len(str(h)) if h is not None else 4
        for r in rows:
            if c - 1 < len(r) and r[c - 1] is not None:
                max_len = max(max_len, len(str(r[c - 1])))
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)
    ws.row_dimensions[1].height = 22

    wb.save(path)


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("LMS 입과명단 분반기")
        root.geometry("820x680")

        self.files: list[str] = []

        # ── 입력 파일 영역 ──────────────────────────────
        f1 = ttk.LabelFrame(root, text="① 입력 엑셀 파일")
        f1.pack(fill="x", padx=10, pady=(10, 6))

        btns = ttk.Frame(f1)
        btns.pack(fill="x", padx=6, pady=4)
        ttk.Button(btns, text="파일 추가...", command=self.add_files).pack(side="left")
        ttk.Button(btns, text="폴더 추가...", command=self.add_folder).pack(side="left", padx=4)
        ttk.Button(btns, text="선택 제거", command=self.remove_selected).pack(side="left")
        ttk.Button(btns, text="전체 비우기", command=self.clear_files).pack(side="left", padx=4)
        ttk.Button(btns, text="헤더 미리보기", command=self.preview_headers).pack(side="left", padx=4)

        list_frame = ttk.Frame(f1)
        list_frame.pack(fill="x", padx=6, pady=(0, 6))
        self.listbox = tk.Listbox(list_frame, height=8, selectmode="extended")
        self.listbox.pack(side="left", fill="x", expand=True)
        sb = ttk.Scrollbar(list_frame, command=self.listbox.yview)
        sb.pack(side="right", fill="y")
        self.listbox.config(yscrollcommand=sb.set)

        # ── 옵션 영역 ──────────────────────────────────
        f2 = ttk.LabelFrame(root, text="② 옵션")
        f2.pack(fill="x", padx=10, pady=6)

        ttk.Label(f2, text="과정 컬럼명:").grid(row=0, column=0, padx=6, pady=6, sticky="e")
        self.col_var = tk.StringVar(value="입과과정")
        ttk.Entry(f2, textvariable=self.col_var, width=22).grid(row=0, column=1, padx=4, pady=6, sticky="w")
        ttk.Label(f2, text="(엑셀 헤더의 과정명 컬럼)", foreground="#666").grid(row=0, column=2, padx=4, sticky="w")

        ttk.Label(f2, text="반 정원:").grid(row=1, column=0, padx=6, pady=6, sticky="e")
        self.size_var = tk.StringVar(value="50")
        ttk.Entry(f2, textvariable=self.size_var, width=8).grid(row=1, column=1, padx=4, pady=6, sticky="w")
        ttk.Label(f2, text="명/반", foreground="#666").grid(row=1, column=2, padx=4, sticky="w")

        ttk.Label(f2, text="출력 폴더:").grid(row=2, column=0, padx=6, pady=6, sticky="e")
        self.out_var = tk.StringVar(value="")
        ttk.Entry(f2, textvariable=self.out_var, width=58).grid(row=2, column=1, columnspan=2, padx=4, pady=6, sticky="we")
        ttk.Button(f2, text="찾아보기...", command=self.choose_out).grid(row=2, column=3, padx=6, pady=6)

        f2.columnconfigure(1, weight=1)

        # ── 실행 ──────────────────────────────────────
        run_frame = ttk.Frame(root)
        run_frame.pack(fill="x", padx=10, pady=4)
        self.run_btn = ttk.Button(run_frame, text="분반 실행", command=self.run)
        self.run_btn.pack(side="left")
        ttk.Button(run_frame, text="출력 폴더 열기", command=self.open_output).pack(side="left", padx=6)

        # ── 로그 ──────────────────────────────────────
        f3 = ttk.LabelFrame(root, text="③ 로그")
        f3.pack(fill="both", expand=True, padx=10, pady=(6, 10))
        self.log_text = scrolledtext.ScrolledText(f3, height=14, font=("Consolas", 9))
        self.log_text.pack(fill="both", expand=True, padx=6, pady=6)

    # ── 파일 리스트 ─────────────────────────────────
    def _add_path(self, path: str):
        if path in self.files:
            return
        self.files.append(path)
        self.listbox.insert("end", f"{os.path.basename(path)}    —  {os.path.dirname(path)}")

    def add_files(self):
        paths = filedialog.askopenfilenames(
            title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        for p in paths:
            self._add_path(p)

    def add_folder(self):
        d = filedialog.askdirectory(title="폴더 선택")
        if not d:
            return
        for f in sorted(os.listdir(d)):
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$") and not f.startswith("_"):
                self._add_path(os.path.join(d, f))

    def remove_selected(self):
        for i in reversed(self.listbox.curselection()):
            del self.files[i]
            self.listbox.delete(i)

    def clear_files(self):
        self.files.clear()
        self.listbox.delete(0, "end")

    def preview_headers(self):
        if not self.files:
            messagebox.showinfo("알림", "먼저 파일을 추가해 주세요.")
            return
        lines = []
        for p in self.files:
            try:
                wb = load_workbook(p, data_only=True, read_only=True)
                ws = wb.active
                row1 = next(ws.iter_rows(values_only=True), [])
                wb.close()
                hdrs = [str(h) if h is not None else "" for h in row1]
                lines.append(f"[{os.path.basename(p)}]  {' | '.join(hdrs)}")
            except Exception as e:
                lines.append(f"[{os.path.basename(p)}]  읽기 실패: {e}")
        self.log("─ 헤더 미리보기 " + "─" * 50)
        for line in lines:
            self.log(line)
        self.log("─" * 65)

    def choose_out(self):
        d = filedialog.askdirectory(title="출력 폴더 선택")
        if d:
            self.out_var.set(d)

    def open_output(self):
        path = self.out_var.get().strip()
        if path and os.path.isdir(path):
            os.startfile(path)
        else:
            messagebox.showinfo("알림", "출력 폴더가 아직 없습니다.")

    # ── 로그 ────────────────────────────────────────
    def log(self, msg: str):
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.update_idletasks()

    # ── 실행 ────────────────────────────────────────
    def run(self):
        if not self.files:
            messagebox.showwarning("알림", "입력 파일을 추가해 주세요.")
            return
        col_name = self.col_var.get().strip()
        if not col_name:
            messagebox.showwarning("알림", "과정 컬럼명을 입력해 주세요.")
            return
        try:
            size = int(self.size_var.get())
            if size < 1:
                raise ValueError
        except ValueError:
            messagebox.showwarning("알림", "반 정원은 양의 정수여야 합니다.")
            return
        out_dir = self.out_var.get().strip()
        if not out_dir:
            out_dir = os.path.join(os.path.dirname(self.files[0]), "_output")
            self.out_var.set(out_dir)
        os.makedirs(out_dir, exist_ok=True)

        self.run_btn.config(state="disabled")
        threading.Thread(target=self._worker, args=(col_name, size, out_dir), daemon=True).start()

    def _worker(self, col_name: str, size: int, out_dir: str):
        try:
            self.log("=" * 65)
            self.log(f"시작 · 컬럼='{col_name}' · 정원={size}명 · 출력={out_dir}")
            self.log("-" * 65)

            canonical_headers = None
            canonical_course_idx = None  # 1-based, canonical 기준
            all_records: list[list] = []  # canonical 컬럼 순서로 정렬된 row 만 저장

            for path in self.files:
                name = os.path.basename(path)
                try:
                    headers, col_idx, data = read_file(path, col_name)
                except Exception as e:
                    self.log(f"  ✗ {name}: 읽기 실패 - {e}")
                    continue
                if headers is None:
                    self.log(f"  ✗ {name}: 빈 파일")
                    continue
                if col_idx is None:
                    shown = " | ".join(str(h) for h in headers if h is not None)
                    self.log(f"  ✗ {name}: '{col_name}' 컬럼 없음. 헤더: {shown}")
                    continue

                if canonical_headers is None:
                    canonical_headers = headers
                    canonical_course_idx = col_idx
                    remapped = data
                elif headers == canonical_headers:
                    remapped = data
                else:
                    idx_map = {}
                    for ci, ch in enumerate(canonical_headers):
                        for fi, fh in enumerate(headers):
                            if fh == ch:
                                idx_map[ci] = fi
                                break
                    remapped = []
                    for row in data:
                        new_row = [row[idx_map[ci]] if ci in idx_map and idx_map[ci] < len(row) else None
                                   for ci in range(len(canonical_headers))]
                        remapped.append(new_row)

                self.log(f"  ✓ {name}: {len(remapped)}행 읽음")
                all_records.extend(remapped)

            if not all_records:
                self.log("처리할 데이터가 없습니다.")
                return

            # 과정별 그룹핑 (입력 순서 유지)
            by_course: "OrderedDict[str, list]" = OrderedDict()
            cidx0 = canonical_course_idx - 1
            for row in all_records:
                val = row[cidx0] if cidx0 < len(row) else None
                key = str(val).strip() if val is not None and str(val).strip() else "(미지정)"
                by_course.setdefault(key, []).append(row)

            self.log("-" * 65)
            self.log(f"총 {len(all_records)}명 · 과정 {len(by_course)}종")
            self.log("-" * 65)

            created = 0
            for course, rows in by_course.items():
                n = len(rows)
                n_classes = max(1, -(-n // size))
                self.log(f"[{course}]  {n}명 → {n_classes}반")
                for ci in range(n_classes):
                    chunk = rows[ci * size:(ci + 1) * size]
                    fname = f"{sanitize_filename(course)}_{ci + 1}반.xlsx"
                    out_path = os.path.join(out_dir, fname)
                    try:
                        write_xlsx(out_path, canonical_headers, chunk)
                        self.log(f"    ✓ {fname}  ({len(chunk)}명)")
                        created += 1
                    except Exception as e:
                        self.log(f"    ✗ {fname} 저장 실패: {e}")

            self.log("-" * 65)
            self.log(f"완료. 파일 {created}개 생성 → {out_dir}")
        finally:
            self.run_btn.config(state="normal")


def main():
    root = tk.Tk()
    try:
        # Windows 기본 폰트
        from tkinter import font as tkfont
        default = tkfont.nametofont("TkDefaultFont")
        default.configure(family="맑은 고딕", size=10)
        root.option_add("*Font", default)
    except Exception:
        pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
